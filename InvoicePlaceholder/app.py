from fastapi import FastAPI, Request
from fastapi.responses import Response
from openpyxl import load_workbook
import xlsxwriter
import os
import re

# For checking if the word is real english word
import enchant


from io import BytesIO
from utils import getMaxConfidence, getAddress, getTotalAmount, getTable

app = FastAPI()


@app.get("/getXLSX")
async def root(request: Request):
    request_body = await request.json()

    model_output = request_body['model_output']

    wb = load_workbook(filename = 'Invoice_template_output_case_study.xlsx')

    ws = wb.active

    en_dictionary = enchant.Dict("en_US")

    _, final_seller_state = getMaxConfidence(model_output, 'SELLER_STATE')
    ws['E3'] = final_seller_state

    _, final_seller_id = getMaxConfidence(model_output, 'SELLER_ID')
    final_seller_id = re.sub(r'[^\w]', '', final_seller_id)
    if final_seller_id != '':
        if(en_dictionary.check(final_seller_id)):
            final_seller_id = ''
    ws['E4'] = final_seller_id

    _, final_seller_name = getMaxConfidence(model_output, 'SELLER_NAME')
    ws['E5'] = final_seller_name

    _, final_seller_gstin_number = getMaxConfidence(
        model_output, 'SELLER_GSTIN_NUMBER')
    final_seller_gstin_number = final_seller_gstin_number.replace('GSTIN', '')
    final_seller_gstin_number = final_seller_gstin_number.replace('gstin', '')
    final_seller_gstin_number = re.sub(r'[^\w]', '', final_seller_gstin_number)
    if final_seller_gstin_number != '':
        if(en_dictionary.check(final_seller_gstin_number)):
            final_seller_gstin_number = ''
    ws['E11'] = final_seller_gstin_number

    _, final_country_of_origin = getMaxConfidence(
        model_output, 'COUNTRY_OF_ORIGIN')
    ws['E12'] = final_country_of_origin

    _, final_currency = getMaxConfidence(model_output, 'CURRENCY')
    ws['E13'] = final_currency

    _, final_invoice_number = getMaxConfidence(model_output, 'INVOICE_NUMBER')
    ws['M3'] = final_invoice_number

    _, final_invoice_date = getMaxConfidence(model_output, 'INVOICE_DATE')
    ws['M4'] = final_invoice_date

    _, final_due_date = getMaxConfidence(model_output, 'DUE_DATE')
    ws['M5'] = final_due_date

    _, final_po_number = getMaxConfidence(model_output, 'PO_NUMBER')
    ws['M10'] = final_po_number

    _, final_buyer_gstin_number = getMaxConfidence(
        model_output, 'BUYER_GSTIN_NUMBER')
    final_buyer_gstin_number = final_buyer_gstin_number.replace('GSTIN', '')
    final_buyer_gstin_number = final_buyer_gstin_number.replace('gstin', '')
    final_buyer_gstin_number = re.sub(r'[^\w]', '', final_buyer_gstin_number)
    if final_buyer_gstin_number != '':
        if(en_dictionary.check(final_buyer_gstin_number)):
            final_buyer_gstin_number = ''
    ws['M13'] = final_buyer_gstin_number

    final_seller_address = getAddress(model_output, 'SELLER_ADDRESS')
    ws['E6'] = final_seller_address

    final_ship_to_address = getAddress(model_output, 'SHIP_TO_ADDRESS')
    ws['M14'] = final_ship_to_address

    final_total_invoice_amount = getTotalAmount(
        model_output, 'TOTAL_INVOICE_AMOUNT_ENTERED_BY_WH_OPERATOR')
    ws['M6'] = final_total_invoice_amount

    final_description = getTotalAmount(model_output, 'DESCRIPTION')
    ws['E14'] = final_description

    table_class_names = ['PRODUCT_ID', 'HSN', 'TITLE', 'QUANTITY', 'UNIT_PRICE',
                         'DISCOUNT_PERCENT', 'SGST_PERCENT', 'CGST_PERCENT', 'IGST_PERCENT', 'TOTAL_AMOUNT']

    final_rows = getTable(model_output, table_class_names)

    row_number = 18
    column = 'B'

    for row in final_rows:
        column = 'B'
        for class_name in table_class_names:
            if class_name in row:
                
                if class_name == 'TOTAL_AMOUNT':
                    row[class_name] = re.sub(r'[^.a-zA-Z0-9 \n\.]', '', row[class_name])
                    try:
                        row[class_name] = int(row[class_name])
                    except:                        
                        try:
                            row[class_name] = float(row[class_name])
                        except:
                            pass

                    ws[column+str(row_number)] = row[class_name]
                    continue

                ws[column+str(row_number)] = row[class_name]
            column = chr(ord(column) + 1)

        row_number += 1

    ws.delete_rows(row_number, 43-row_number)

    ws['K' + str(row_number)] = '=SUM(K18:K' + str(row_number-1)

    

    # workbook.close()
    file_name = request_body['filename']
    file_name = os.path.join('results', file_name + '.xlsx')

    wb.save(file_name)
    print("saved to : ", file_name)

    with open(file_name, 'rb') as f:
        xlsx_file = f.read()

    return Response(content=xlsx_file,
                    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
